import openpyxl
import json

# Read all exercises
wb = openpyxl.load_workbook('exercise/exercises.xlsx')
ws = wb.active
exercises = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    if row[0]:
        exercises.append({'category': row[0], 'bodyPart': row[1], 'name': row[2], 'equipment': row[3]})

# Read all templates
templates = []
for day in range(1, 8):
    wb = openpyxl.load_workbook(f'template/Day {day}.xlsx')
    ws = wb.active
    exs = [row[0] for row in ws.iter_rows(values_only=True) if row[0]]
    templates.append({'id': f'day{day}', 'name': f'Day {day}', 'exercises': exs})

print("=== EXERCISES ===")
print(json.dumps(exercises, indent=2))
print("\n=== TEMPLATES ===")
print(json.dumps(templates, indent=2))

